"""
Template mapper module for mapping processed data to Excel templates.
"""
from typing import Dict, Any, Optional, List
from pathlib import Path
import pandas as pd
import logging

logger = logging.getLogger(__name__)


class TemplateMappingError(Exception):
    """Custom exception for template mapping errors."""
    pass


class TemplateMapper:
    """Mapper for converting processed data to Excel template format."""
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        Initialize the template mapper.
        
        Args:
            config: Configuration dictionary containing template mapping rules
        """
        self.config = config or self._get_default_config()
        self.column_mapping = self.config.get('template_mapping', {}).get('columns', {})
    
    def map_to_template(self, 
                       layout_data: Dict[str, Dict[str, Any]], 
                       template_path: Optional[Path] = None,
                       output_path: Optional[Path] = None) -> Path:
        """
        Map processed layout data to Excel template format.
        
        Args:
            layout_data: Processed data from RuleEngine
            template_path: Path to Excel template (optional)
            output_path: Path for output Excel file
            
        Returns:
            Path to the generated Excel file
            
        Raises:
            TemplateMappingError: If mapping fails
        """
        try:
            # Convert layout data to DataFrame format
            df_data = self._convert_to_dataframe_format(layout_data)
            
            # Create DataFrame
            df = pd.DataFrame(df_data)
            
            # Sort by priority and then by net name
            df = df.sort_values(['priority', 'net_name'])
            
            # Reorder columns according to template mapping
            df = self._reorder_columns(df)
            
            # Determine output path
            if output_path is None:
                output_path = Path.cwd() / "layout_guide_output.xlsx"
            
            # Save to Excel
            self._save_to_excel(df, output_path)
            
            logger.info(f"Successfully mapped {len(layout_data)} nets to Excel template: {output_path}")
            return output_path
            
        except Exception as e:
            raise TemplateMappingError(f"Failed to map data to template: {str(e)}")
    
    def _convert_to_dataframe_format(self, layout_data: Dict[str, Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Convert layout data to format suitable for DataFrame.
        
        Args:
            layout_data: Layout data from RuleEngine
            
        Returns:
            List of dictionaries for DataFrame creation
        """
        df_data = []
        
        for net_name, net_info in layout_data.items():
            row_data = {
                'Category': net_info.get('category', 'Other'),
                'Net Name': net_name,
                'Pin (MT7921)': net_info.get('pin_assignment', 'TBD'),
                'Description': net_info.get('description', ''),
                'Impedance': net_info.get('impedance', '50 Ohm'),
                'Type': net_info.get('signal_type', 'Single-End'),
                'Width': net_info.get('width', 'TBD'),
                'Length Limit (mil)': net_info.get('length_limit', 'TBD'),
                'Spacing': net_info.get('spacing', 'TBD'),
                'Shielding': net_info.get('shielding', 'Optional'),
                'Layer Stack': net_info.get('layer_stack', 'Any'),
                'Notes': net_info.get('notes', ''),
                'priority': net_info.get('priority', 999)  # For sorting, will be removed
            }
            df_data.append(row_data)
        
        return df_data
    
    def _reorder_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Reorder DataFrame columns according to template mapping.
        
        Args:
            df: Input DataFrame
            
        Returns:
            DataFrame with reordered columns
        """
        # Define desired column order
        desired_order = [
            'Category',
            'Net Name', 
            'Pin (MT7921)',
            'Description',
            'Impedance',
            'Type',
            'Width',
            'Length Limit (mil)',
            'Spacing',
            'Shielding',
            'Layer Stack',
            'Notes'
        ]
        
        # Keep only columns that exist in the DataFrame
        available_cols = [col for col in desired_order if col in df.columns]
        
        # Return DataFrame with reordered columns
        return df[available_cols]
    
    def _save_to_excel(self, df: pd.DataFrame, output_path: Path) -> None:
        """
        Save DataFrame to Excel file with formatting.
        
        Args:
            df: DataFrame to save
            output_path: Path for output file
        """
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Layout Guide', index=False)
                
                # Get the workbook and worksheet for formatting
                workbook = writer.book
                worksheet = writer.sheets['Layout Guide']
                
                # Apply basic formatting
                self._apply_excel_formatting(worksheet, len(df))
                
        except Exception as e:
            logger.error(f"Error saving Excel file: {e}")
            # Fallback to basic save without formatting
            df.to_excel(output_path, index=False)
    
    def _apply_excel_formatting(self, worksheet, row_count: int) -> None:
        """
        Apply basic formatting to Excel worksheet.
        
        Args:
            worksheet: openpyxl worksheet object
            row_count: Number of data rows
        """
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # Header formatting
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Auto-adjust column widths
            for col in range(1, worksheet.max_column + 1):
                column_letter = get_column_letter(col)
                max_length = 0
                
                for row in range(1, min(row_count + 2, 100)):  # Limit check to first 100 rows
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                
                # Set column width with reasonable limits
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
        except ImportError:
            logger.warning("openpyxl.styles not available, skipping Excel formatting")
        except Exception as e:
            logger.warning(f"Error applying Excel formatting: {e}")
    
    def validate_template(self, template_path: Path) -> bool:
        """
        Validate if the Excel template has expected structure.
        
        Args:
            template_path: Path to template file
            
        Returns:
            True if template is valid
        """
        try:
            if not template_path.exists():
                logger.error(f"Template file not found: {template_path}")
                return False
            
            # Try to read the template
            df = pd.read_excel(template_path)
            
            # Check if template has minimum required columns
            required_columns = ['Category', 'Net Name', 'Description']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                logger.error(f"Template missing required columns: {missing_columns}")
                return False
            
            logger.info(f"Template validation successful: {template_path}")
            return True
            
        except Exception as e:
            logger.error(f"Template validation failed: {e}")
            return False
    
    def get_template_info(self, template_path: Path) -> Dict[str, Any]:
        """
        Get information about a template file.
        
        Args:
            template_path: Path to template file
            
        Returns:
            Dictionary with template information
        """
        try:
            df = pd.read_excel(template_path)
            return {
                'columns': list(df.columns),
                'row_count': len(df),
                'file_size': template_path.stat().st_size,
                'valid': self.validate_template(template_path)
            }
        except Exception as e:
            return {'error': str(e), 'valid': False}
    
    def _get_default_config(self) -> Dict[str, Any]:
        """Get default template mapping configuration."""
        return {
            'template_mapping': {
                'columns': {
                    'Category': 'Category',
                    'Net Name': 'Net Name',
                    'Pin': 'Pin (MT7921)',
                    'Description': 'Description',
                    'Impedance': 'Impedance',
                    'Type': 'Type',
                    'Width': 'Width',
                    'Length Limit': 'Length Limit (mil)',
                    'Spacing': 'Spacing',
                    'Shielding': 'Shielding',
                    'Layer Stack': 'Layer Stack',
                    'Notes': 'Notes'
                }
            }
        }